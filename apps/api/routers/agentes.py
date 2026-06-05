"""
Router del catÃ¡logo de agentes de seguros.

GET    /agentes                             â€” listado con bÃºsqueda y filtros
POST   /agentes                             â€” crear agente (directores + gerentes)
GET    /agentes/{id}                        â€” perfil completo con telefonos, emails, asistentes
PATCH  /agentes/{id}                        â€” actualizar datos del agente
DELETE /agentes/{id}                        â€” soft-delete

POST   /agentes/{id}/telefonos              â€” agregar telÃ©fono
PATCH  /agentes/{id}/telefonos/{tel_id}/preferente â€” marcar como preferente
DELETE /agentes/{id}/telefonos/{tel_id}     â€” eliminar telÃ©fono

POST   /agentes/{id}/emails                 â€” agregar email
PATCH  /agentes/{id}/emails/{email_id}/preferente â€” marcar como preferente
DELETE /agentes/{id}/emails/{email_id}      â€” eliminar email

POST   /agentes/{id}/asistentes             â€” agregar asistente
PATCH  /agentes/{id}/asistentes/{asist_id}  â€” actualizar asistente
"""

import contextlib
from datetime import date
from io import BytesIO
from uuid import UUID

import structlog
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from supabase import Client

from core.auth import get_current_user, require_roles
from core.busqueda import filtro_busqueda_or
from core.database import get_admin_db, get_db
from models.agente import (
    AgenteCreate,
    AgenteEmailCreate,
    AgenteEmailResponse,
    AgenteImportRow,
    AgenteListItem,
    AgenteResponse,
    AgenteUpdate,
    AsistenteCreate,
    AsistenteResponse,
    AsistenteUpdate,
    ImportResponse,
    ImportResultItem,
    TelefonoCreate,
    TelefonoResponse,
    TipoTelefono,
)
from models.usuario import RolUsuario, UsuarioToken

log = structlog.get_logger(__name__)
router = APIRouter(prefix="/agentes", tags=["agentes"])

_ESCRITURA = [
    Depends(require_roles(RolUsuario.director_general, RolUsuario.director_ops, RolUsuario.gerente))
]


def _check_agente_existe(db: Client, agente_id: UUID) -> dict:
    result = (
        db.table("agente").select("id, activo").eq("id", str(agente_id)).maybe_single().execute()
    )
    if not result.data:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Agente no encontrado.")
    return result.data


# ---------------------------------------------------------------------------
# GET /agentes
# ---------------------------------------------------------------------------


@router.get("", response_model=list[AgenteListItem])
def listar_agentes(
    q: str | None = Query(default=None, description="BÃºsqueda por nombre o CUA"),
    activo: bool | None = Query(default=True),
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
    db: Client = Depends(get_db),
) -> list[AgenteListItem]:
    """Lista de agentes con bÃºsqueda y filtros. Visible para todos los roles."""
    query = db.table("agente").select(
        "id, cua, nombre, nombre_comercial, rfc, fecha_afiliacion, activo, "
        "agente_email!left(email, preferente), "
        "agente_telefono!left(numero, preferente)"
    )

    if activo is not None:
        query = query.eq("activo", activo)

    if q:
        query = query.or_(filtro_busqueda_or(q, "nombre", "cua"))

    result = query.order("nombre").range(offset, offset + limit - 1).execute()

    items = []
    for row in result.data:
        emails = row.pop("agente_email", []) or []
        telefonos = row.pop("agente_telefono", []) or []

        email_pref = next((e["email"] for e in emails if e.get("preferente")), None)
        tel_pref = next((t["numero"] for t in telefonos if t.get("preferente")), None)

        items.append(
            AgenteListItem(
                **row,
                email_preferente=email_pref,
                telefono_preferente=tel_pref,
            )
        )

    return items


# ---------------------------------------------------------------------------
# POST /agentes
# ---------------------------------------------------------------------------


@router.post(
    "", response_model=AgenteResponse, status_code=status.HTTP_201_CREATED, dependencies=_ESCRITURA
)
def crear_agente(
    body: AgenteCreate,
    db: Client = Depends(get_db),
) -> AgenteResponse:
    try:
        result = (
            db.table("agente")
            .insert(body.model_dump(mode="json", exclude_none=True))
            .select(
                "id, cua, nombre, nombre_comercial, rfc, fecha_afiliacion, notas, activo, created_at, updated_at"
            )
            .execute()
        )
        if result.data:
            result.data = result.data[0]
    except Exception as exc:
        if "uq_agente_cua" in str(exc):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Ya existe un agente con CUA '{body.cua}'.",
            ) from exc
        if "uq_agente_rfc" in str(exc):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Ya existe un agente con RFC '{body.rfc}'.",
            ) from exc
        raise exc from None

    data = result.data
    data.update({"telefonos": [], "emails": [], "asistentes": []})
    log.info("agente_creado", id=data["id"], cua=body.cua)
    return AgenteResponse.model_validate(data)


# ---------------------------------------------------------------------------
# GET /agentes/template — descarga plana Excel con columnas requeridas
# ---------------------------------------------------------------------------


@router.get("/template", dependencies=[Depends(get_current_user)])
def descargar_template():
    """Genera un archivo Excel en blanco con los encabezados correctos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Agentes"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "CUA *",
        "Nombre completo *",
        "Nombre comercial",
        "RFC",
        "Fecha de afiliación (YYYY-MM-DD)",
        "Correo electrónico",
        "Teléfono",
        "Tipo de teléfono (celular/oficina/casa/whatsapp/otro)",
        "Notas",
    ]

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    example_rows = [
        [
            "A000123456",
            "Juan Pérez García",
            "Agencia JP",
            "PEGJ800101ABC",
            "2020-01-15",
            "juan@email.com",
            "55 1234 5678",
            "celular",
            "Agente con 5 años de experiencia",
        ],
        [
            "A000234567",
            "María López Hernández",
            "Seguros ML",
            "LOPM850202XYZ",
            "2019-06-01",
            "maria@email.com",
            "55 8765 4321",
            "whatsapp",
            "",
        ],
    ]

    example_fill = PatternFill("solid", fgColor="F8FAFC")
    for r, row_data in enumerate(example_rows, start=2):
        for c, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = example_fill
            cell.border = thin_border
            if c == 1:
                cell.font = Font(bold=True)

    col_widths = [20, 35, 25, 15, 28, 30, 18, 30, 40]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=olimpo_agentes_template.xlsx",
            "Content-Length": str(buffer.getbuffer().nbytes),
        },
    )


# ---------------------------------------------------------------------------
# GET /agentes/{id}
# ---------------------------------------------------------------------------


@router.get("/{agente_id}", response_model=AgenteResponse)
def obtener_agente(
    agente_id: UUID,
    db: Client = Depends(get_db),
) -> AgenteResponse:
    result = (
        db.table("agente")
        .select(
            "id, cua, nombre, nombre_comercial, rfc, fecha_afiliacion, notas, activo, created_at, updated_at, "
            "agente_telefono!left(id, agente_id, tipo, numero, preferente, created_at), "
            "agente_email!left(id, agente_id, email, preferente, created_at), "
            "asistente!asistente_agente_id_fkey!left(id, agente_id, nombre, email, telefono, activo, created_at, updated_at)"
        )
        .eq("id", str(agente_id))
        .maybe_single()
        .execute()
    )

    if not result.data:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Agente no encontrado.")

    data = result.data
    # Renombrar claves de join anidado a los nombres del modelo
    data["telefonos"] = data.pop("agente_telefono", []) or []
    data["emails"] = data.pop("agente_email", []) or []
    data["asistentes"] = data.pop("asistente", []) or []

    return AgenteResponse.model_validate(data)


# ---------------------------------------------------------------------------
# PATCH /agentes/{id}
# ---------------------------------------------------------------------------


@router.patch("/{agente_id}", response_model=AgenteResponse, dependencies=_ESCRITURA)
def actualizar_agente(
    agente_id: UUID,
    body: AgenteUpdate,
    db: Client = Depends(get_db),
) -> AgenteResponse:
    cambios = body.model_dump(exclude_none=True)
    if not cambios:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail="No se enviaron campos para actualizar.",
        )

    _check_agente_existe(db, agente_id)

    try:
        db.table("agente").update(cambios).eq("id", str(agente_id)).execute()
    except Exception as exc:
        if "uq_agente_rfc" in str(exc):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT, detail="RFC ya registrado en otro agente."
            ) from exc
        raise exc from None

    return obtener_agente(agente_id, db)


# ---------------------------------------------------------------------------
# DELETE /agentes/{id} â€” soft-delete
# ---------------------------------------------------------------------------


@router.delete("/{agente_id}", status_code=status.HTTP_204_NO_CONTENT, dependencies=_ESCRITURA)
def desactivar_agente(
    agente_id: UUID,
    usuario: UsuarioToken = Depends(get_current_user),
    db: Client = Depends(get_db),
) -> None:
    _check_agente_existe(db, agente_id)
    db.table("agente").update({"activo": False}).eq("id", str(agente_id)).execute()
    log.info("agente_desactivado", id=str(agente_id), por=str(usuario.id))


# ---------------------------------------------------------------------------
# TelÃ©fonos
# ---------------------------------------------------------------------------


@router.post(
    "/{agente_id}/telefonos",
    response_model=TelefonoResponse,
    status_code=status.HTTP_201_CREATED,
    dependencies=_ESCRITURA,
)
def agregar_telefono(
    agente_id: UUID,
    body: TelefonoCreate,
    db: Client = Depends(get_db),
) -> TelefonoResponse:
    _check_agente_existe(db, agente_id)

    payload = body.model_dump()
    payload["agente_id"] = str(agente_id)

    result = db.table("agente_telefono").insert(payload).select("*").execute()
    if result.data:
        result.data = result.data[0]
    return TelefonoResponse.model_validate(result.data)


@router.patch(
    "/{agente_id}/telefonos/{telefono_id}/preferente",
    response_model=TelefonoResponse,
    dependencies=_ESCRITURA,
)
def marcar_telefono_preferente(
    agente_id: UUID,
    telefono_id: UUID,
    db: Client = Depends(get_db),
) -> TelefonoResponse:
    """
    Marca este telÃ©fono como preferente y quita el preferente anterior.
    El Ã­ndice Ãºnico parcial uq_agente_telefono_preferente garantiza solo uno activo.
    Se hace en dos pasos para evitar conflicto de constraint.
    """
    # Quitar preferente actual
    db.table("agente_telefono").update({"preferente": False}).eq("agente_id", str(agente_id)).eq(
        "preferente", True
    ).execute()

    # Marcar el nuevo
    result = (
        db.table("agente_telefono")
        .update({"preferente": True})
        .eq("id", str(telefono_id))
        .eq("agente_id", str(agente_id))
        .select("*")
        .execute()
    )
    if result.data:
        result.data = result.data[0]

    if not result.data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="TelÃ©fono no encontrado."
        )
    return TelefonoResponse.model_validate(result.data)


@router.delete(
    "/{agente_id}/telefonos/{telefono_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=_ESCRITURA,
)
def eliminar_telefono(
    agente_id: UUID,
    telefono_id: UUID,
    db: Client = Depends(get_db),
) -> None:
    db.table("agente_telefono").delete().eq("id", str(telefono_id)).eq(
        "agente_id", str(agente_id)
    ).execute()


# ---------------------------------------------------------------------------
# Emails del agente
# ---------------------------------------------------------------------------


@router.post(
    "/{agente_id}/emails",
    response_model=AgenteEmailResponse,
    status_code=status.HTTP_201_CREATED,
    dependencies=_ESCRITURA,
)
def agregar_email(
    agente_id: UUID,
    body: AgenteEmailCreate,
    db: Client = Depends(get_db),
) -> AgenteEmailResponse:
    _check_agente_existe(db, agente_id)

    payload = body.model_dump()
    payload["agente_id"] = str(agente_id)

    try:
        result = db.table("agente_email").insert(payload).select("*").execute()
        if result.data:
            result.data = result.data[0]
    except Exception as exc:
        msg = str(exc)
        if "uq_agente_email_email" in msg or "duplicate" in msg.lower():
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"El correo '{body.email}' ya estÃ¡ registrado.",
            ) from exc
        if "ya estÃ¡ registrado como email de un asistente" in msg:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=str(exc)) from exc
        raise exc from None

    return AgenteEmailResponse.model_validate(result.data)


@router.patch(
    "/{agente_id}/emails/{email_id}/preferente",
    response_model=AgenteEmailResponse,
    dependencies=_ESCRITURA,
)
def marcar_email_preferente(
    agente_id: UUID,
    email_id: UUID,
    db: Client = Depends(get_db),
) -> AgenteEmailResponse:
    db.table("agente_email").update({"preferente": False}).eq("agente_id", str(agente_id)).eq(
        "preferente", True
    ).execute()

    result = (
        db.table("agente_email")
        .update({"preferente": True})
        .eq("id", str(email_id))
        .eq("agente_id", str(agente_id))
        .select("*")
        .execute()
    )
    if result.data:
        result.data = result.data[0]

    if not result.data:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Email no encontrado.")
    return AgenteEmailResponse.model_validate(result.data)


@router.delete(
    "/{agente_id}/emails/{email_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=_ESCRITURA,
)
def eliminar_email(
    agente_id: UUID,
    email_id: UUID,
    db: Client = Depends(get_db),
) -> None:
    db.table("agente_email").delete().eq("id", str(email_id)).eq(
        "agente_id", str(agente_id)
    ).execute()


# ---------------------------------------------------------------------------
# Asistentes
# ---------------------------------------------------------------------------


@router.post(
    "/{agente_id}/asistentes",
    response_model=AsistenteResponse,
    status_code=status.HTTP_201_CREATED,
    dependencies=_ESCRITURA,
)
def agregar_asistente(
    agente_id: UUID,
    body: AsistenteCreate,
    db: Client = Depends(get_db),
) -> AsistenteResponse:
    _check_agente_existe(db, agente_id)

    payload = body.model_dump(exclude_none=True)
    payload["agente_id"] = str(agente_id)

    try:
        result = db.table("asistente").insert(payload).select("*").execute()
        if result.data:
            result.data = result.data[0]
    except Exception as exc:
        msg = str(exc)
        if "uq_asistente_email" in msg or "duplicate" in msg.lower():
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"El correo '{body.email}' ya estÃ¡ registrado como asistente.",
            ) from exc
        if "ya estÃ¡ registrado como email de un agente" in msg:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=str(exc)) from exc
        raise exc from None

    return AsistenteResponse.model_validate(result.data)


@router.patch(
    "/{agente_id}/asistentes/{asistente_id}",
    response_model=AsistenteResponse,
    dependencies=_ESCRITURA,
)
def actualizar_asistente(
    agente_id: UUID,
    asistente_id: UUID,
    body: AsistenteUpdate,
    db: Client = Depends(get_db),
) -> AsistenteResponse:
    cambios = body.model_dump(exclude_none=True)
    if not cambios:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="No se enviaron campos."
        )

    result = (
        db.table("asistente")
        .update(cambios)
        .eq("id", str(asistente_id))
        .eq("agente_id", str(agente_id))
        .select("*")
        .execute()
    )
    if result.data:
        result.data = result.data[0]

    if not result.data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Asistente no encontrado."
        )
    return AsistenteResponse.model_validate(result.data)


@router.delete(
    "/{agente_id}/asistentes/{asistente_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=_ESCRITURA,
)
def eliminar_asistente(
    agente_id: UUID,
    asistente_id: UUID,
    db: Client = Depends(get_db),
) -> None:
    db.table("asistente").delete().eq("id", str(asistente_id)).eq(
        "agente_id", str(agente_id)
    ).execute()


# ---------------------------------------------------------------------------
# POST /agentes/import
# ---------------------------------------------------------------------------


@router.post(
    "/import",
    response_model=ImportResponse,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(get_current_user)] + _ESCRITURA,
)
async def importar_agentes(
    file: UploadFile = File(...),
    db: Client = Depends(get_db),
):
    """
    Recibe un archivo Excel (.xlsx), valida cada fila contra AgenteImportRow,
    detecta CUA duplicados en el propio archivo, consulta cuáles ya existen
    en la DB (por CUA) y hace INSERT en batch o individual con errores capturados.

    Retorna el detalle de cada fila: éxito, id generado o error.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail="Formato no soportado. Usa un archivo .xlsx o .xls",
        )

    contents = await file.read()
    try:
        wb = load_workbook(BytesIO(contents), data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="No se pudo leer el archivo Excel.",
        ) from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="El archivo no contiene filas de datos.",
        )

    header_row = [str(c).strip() if c is not None else "" for c in rows[0]]
    col_map = {h.lower(): i for i, h in enumerate(header_row)}

    required = {"cua": None, "nombre completo": None}
    for key in required:
        if key not in col_map:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Falta la columna requerida: '{key}'",
            )

    import_rows: list[AgenteImportRow] = []
    preview_errors: list[tuple[int, str]] = []

    for idx, row in enumerate(rows[1:], start=2):

        def _col(key: str, default=None):
            return row[col_map[key]] if key in col_map and col_map[key] < len(row) else default

        raw = AgenteImportRow(
            cua=str(_col("cua") or "").strip(),
            nombre=str(_col("nombre completo") or "").strip(),
            nombre_comercial=str(_col("nombre comercial") or "").strip() or None,
            rfc=str(_col("rfc") or "").strip() or None,
            fecha_afiliacion=str(_col("fecha de afiliación (yyyy-mm-dd)") or "").strip() or None,
            email=str(_col("correo electrónico") or "").strip() or None,
            telefono=str(_col("teléfono") or "").strip() or None,
            tipo_telefono=str(
                _col("tipo de teléfono (celular/oficina/casa/whatsapp/otro)") or ""
            ).strip()
            or None,
            notas=str(_col("notas") or "").strip() or None,
        )

        row_errors = []
        if not raw.cua:
            row_errors.append("CUA es requerido")
        if not raw.nombre or len(raw.nombre) < 2:
            row_errors.append("Nombre es requerido (mínimo 2 caracteres)")

        if raw.fecha_afiliacion:
            try:
                date.fromisoformat(raw.fecha_afiliacion)
            except ValueError:
                row_errors.append(
                    f"Fecha de afiliación inválida: '{raw.fecha_afiliacion}' — usar formato YYYY-MM-DD"
                )
                raw.fecha_afiliacion = None

        if raw.tipo_telefono and raw.tipo_telefono not in {t.value for t in TipoTelefono}:
            row_errors.append(f"Tipo de teléfono desconocido: '{raw.tipo_telefono}'")

        if row_errors:
            preview_errors.append((idx, "; ".join(row_errors)))
        else:
            import_rows.append(raw)

    if not import_rows:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="No se encontraron filas válidas para importar. "
            + "; ".join(f"Fila {r}: {e}" for r, e in preview_errors),
        )

    cuas = [r.cua for r in import_rows]
    if len(cuas) != len(set(cuas)):
        dupes = [c for c in cuas if cuas.count(c) > 1]
        seen: set[str] = set()
        for d in dupes:
            if d not in seen:
                seen.add(d)
                preview_errors.append((-1, f"CUA duplicado en el archivo: '{d}'"))
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="El archivo contiene CUA duplicados: " + ", ".join(sorted(seen)),
        )

    admin_db = get_admin_db()

    existing: dict[str, dict] = {}
    if cuas:
        res = admin_db.table("agente").select("id, cua").in_("cua", cuas).execute()
        for r in res.data or []:
            existing[r["cua"]] = r

    results: list[ImportResultItem] = []
    exitosos = 0
    fallidos = 0
    errores_duplicados = 0

    for idx, row_data in enumerate(import_rows, start=2):
        cua = row_data.cua

        if cua in existing:
            errores_duplicados += 1
            results.append(
                ImportResultItem(
                    row=idx, cua=cua, success=False, error="Ya existe un agente con este CUA"
                )
            )
            continue

        try:
            insert_payload: dict = {
                "cua": cua,
                "nombre": row_data.nombre,
                "nombre_comercial": row_data.nombre_comercial,
                "rfc": row_data.rfc,
                "notas": row_data.notas,
            }
            if row_data.fecha_afiliacion:
                insert_payload["fecha_afiliacion"] = row_data.fecha_afiliacion

            ins = admin_db.table("agente").insert(insert_payload).execute()
            if not ins.data:
                raise Exception("No se devolvió id")

            created = ins.data[0]
            agente_id: str = created["id"]

            if row_data.email:
                with contextlib.suppress(Exception):
                    admin_db.table("agente_email").insert(
                        {
                            "agente_id": agente_id,
                            "email": row_data.email,
                            "preferente": True,
                        }
                    ).execute()

            if row_data.telefono:
                with contextlib.suppress(Exception):
                    admin_db.table("agente_telefono").insert(
                        {
                            "agente_id": agente_id,
                            "numero": row_data.telefono,
                            "tipo": row_data.tipo_telefono or TipoTelefono.celular.value,
                            "preferente": True,
                        }
                    ).execute()

            exitosos += 1
            results.append(ImportResultItem(row=idx, cua=cua, success=True, agente_id=agente_id))
            existing[cua] = {"id": agente_id}

        except Exception as exc:
            fallidos += 1
            results.append(ImportResultItem(row=idx, cua=cua, success=False, error=str(exc)))

    detalle_parts = []
    if exitosos:
        detalle_parts.append(f"{exitosos} agente(s) importado(s) correctamente")
    if errores_duplicados:
        detalle_parts.append(f"{errores_duplicados} skipped por CUA duplicado")
    if fallidos:
        detalle_parts.append(f"{fallidos} fallido(s)")
    detalle = "; ".join(detalle_parts) or "Sin resultados"

    log.info(
        "agentes_importados",
        total=len(import_rows),
        exitosos=exitosos,
        fallidos=fallidos,
        errores_duplicados=errores_duplicados,
    )

    return ImportResponse(
        total=len(import_rows),
        exitosos=exitosos,
        fallidos=fallidos,
        errores_duplicados=errores_duplicados,
        results=results,
        detalle=detalle,
    )
